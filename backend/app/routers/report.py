import csv
import io
import re
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import FileResponse

import hashlib
from pathlib import Path
import xml.etree.ElementTree as ET
import zipfile
from uuid import uuid4

from ..config import (
    DEFAULT_TEMPLATE_NAME,
    INSTRUMENT_CATALOG_AUTO_DIR,
    INSTRUMENT_CATALOG_AUTO_ENABLED,
    INSTRUMENT_CATALOG_AUTO_KEYWORDS,
    LOCAL_DOCUMENT_LIBRARY_FILE,
    OUTPUT_DIR,
    TEMPLATE_DIR,
    UPLOAD_DIR,
)
from ..schemas.device_report import (
    DeviceFields,
    EditorPrefillRequest,
    EditorPrefillResponse,
    ExcelBatchRequest,
    ExcelBatchResponse,
    ExcelInspectRequest,
    ExcelInspectResponse,
    ExcelPreviewRequest,
    ExcelPreviewResponse,
    ExtractRequest,
    ReportValidation,
    ReportRequest,
    ReportResponse,
    TemplateMatchRequest,
    TemplateMatchResponse,
    InstrumentCatalogParseResponse,
)
from ..services.excel_batch_service import inspect_excel_records, parse_excel_rows, preview_excel_sheet
from ..services.extract_service import extract_fields
from ..services.local_document_library_service import load_local_document_library, rebuild_local_document_library
from ..services.template_mapping_library_service import get_editor_schema
from ..services.template_service import (
    FixedTemplateFillError,
    get_template_editor_prefill,
    list_available_templates,
    match_template_name,
    render_report,
)

router = APIRouter()


@router.post("/extract", response_model=DeviceFields)
def extract(request: ExtractRequest) -> DeviceFields:
    fields = extract_fields(request.raw_text)
    return DeviceFields(**fields)


@router.post("/report", response_model=ReportResponse)
def create_report(request: ReportRequest) -> ReportResponse:
    template_name = request.template_name or DEFAULT_TEMPLATE_NAME
    context = request.fields.model_dump()
    source_file_path = _find_uploaded_file(request.source_file_id) if request.source_file_id else None
    try:
        report_id, output_path = render_report(
            template_name=template_name,
            context=context,
            source_file_path=source_file_path,
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except FixedTemplateFillError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    output_format = output_path.suffix.lower().lstrip(".")
    preview_url = f"/api/report/{report_id}/view" if output_format == "html" else None

    return ReportResponse(
        report_id=report_id,
        download_url=f"/api/report/{report_id}/download",
        preview_url=preview_url,
        output_format=output_format,
        validation=_build_report_validation(output_path),
    )


@router.post("/report/batch-from-excel", response_model=ExcelBatchResponse)
def create_report_batch_from_excel(request: ExcelBatchRequest) -> ExcelBatchResponse:
    file_path = _find_uploaded_file(request.file_id)
    if not file_path:
        raise HTTPException(status_code=404, detail="File not found")
    if file_path.suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Only .xlsx is supported for batch mode")

    rows, errors = parse_excel_rows(
        file_path=file_path,
        sheet_name=request.sheet_name,
        default_template_name=request.default_template_name or "",
    )
    if not rows:
        raise HTTPException(status_code=422, detail="Excel has no valid rows for generation")

    batch_id = uuid4().hex
    zip_path = OUTPUT_DIR / f"{batch_id}__excel_batch.zip"
    total_rows = len(rows) + len(errors)
    generated_count = 0
    skipped_count = len(errors)

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for row in rows:
            try:
                _, output_path = render_report(
                    template_name=row["template_name"],
                    context=row["fields"],
                    source_file_path=file_path,
                )
                arc_name = f"{row['row_number']:04d}_{row['row_name']}__{_get_report_download_name(output_path)}"
                zf.write(output_path, arcname=arc_name)
                generated_count += 1
            except Exception as exc:
                skipped_count += 1
                errors.append(f"第 {row['row_number']} 行生成失败：{str(exc)}")

    if generated_count == 0:
        raise HTTPException(status_code=422, detail="Excel rows could not be generated")

    return ExcelBatchResponse(
        batch_id=batch_id,
        download_url=f"/api/report/batch/{batch_id}/download",
        total_rows=total_rows,
        generated_count=generated_count,
        skipped_count=skipped_count,
        errors=errors[:200],
    )


@router.post("/report/inspect-from-excel", response_model=ExcelInspectResponse)
def inspect_report_batch_from_excel(request: ExcelInspectRequest) -> ExcelInspectResponse:
    file_path = _find_uploaded_file(request.file_id)
    if not file_path:
        raise HTTPException(status_code=404, detail="File not found")
    if file_path.suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Only .xlsx is supported for batch mode")

    rows, errors = inspect_excel_records(
        file_path=file_path,
        sheet_name=request.sheet_name,
        default_template_name=request.default_template_name or "",
    )
    total_rows = len(rows)
    valid_rows = len([row for row in rows if not row.get("error")])
    skipped_rows = len(rows) - valid_rows
    records = [
        {
            "sheet_name": row.get("sheet_name", ""),
            "row_number": row.get("row_number", 0),
            "row_name": row.get("row_name", ""),
            "template_name": row.get("template_name", ""),
            "fields": row.get("fields", {}),
            "error": row.get("error", ""),
        }
        for row in rows
    ]
    return ExcelInspectResponse(
        total_rows=total_rows,
        valid_rows=valid_rows,
        skipped_rows=skipped_rows,
        errors=errors[:200],
        records=records,
    )


@router.post("/report/preview-from-excel", response_model=ExcelPreviewResponse)
def preview_report_batch_from_excel(request: ExcelPreviewRequest) -> ExcelPreviewResponse:
    file_path = _find_uploaded_file(request.file_id)
    if not file_path:
        raise HTTPException(status_code=404, detail="File not found")
    if file_path.suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Only .xlsx is supported for preview mode")

    payload = preview_excel_sheet(file_path=file_path, sheet_name=request.sheet_name)
    return ExcelPreviewResponse(
        sheet_names=payload.get("sheet_names", []),
        sheet_name=str(payload.get("sheet_name", "") or ""),
        title=str(payload.get("title", "") or ""),
        headers=payload.get("headers", []),
        rows=payload.get("rows", []),
        row_numbers=payload.get("row_numbers", []),
        total_rows=payload.get("total_rows", 0),
        truncated=bool(payload.get("truncated", False)),
    )


@router.get("/report/{report_id}/download")
def download_report(report_id: str) -> FileResponse:
    file_path = _find_report_file(report_id)
    if not file_path:
        raise HTTPException(status_code=404, detail="Report not found")
    return FileResponse(
        path=str(file_path),
        media_type=_guess_media_type(file_path),
        filename=_get_report_download_name(file_path),
    )


@router.get("/report/batch/{batch_id}/download")
def download_report_batch(batch_id: str) -> FileResponse:
    file_path = _find_report_batch_zip(batch_id)
    if not file_path:
        raise HTTPException(status_code=404, detail="Batch report not found")
    return FileResponse(
        path=str(file_path),
        media_type="application/zip",
        filename=_get_report_download_name(file_path),
    )


@router.get("/report/{report_id}/view")
def view_report(report_id: str) -> FileResponse:
    file_path = _find_report_file(report_id)
    if not file_path:
        raise HTTPException(status_code=404, detail="Report not found")
    if file_path.suffix.lower() != ".html":
        raise HTTPException(status_code=400, detail="Preview only supports html report")
    return FileResponse(path=str(file_path), media_type="text/html")


@router.get("/templates")
def list_templates() -> dict[str, list[str]]:
    return {"templates": list_available_templates()}


@router.get("/templates/download")
def download_template(template_name: str) -> FileResponse:
    available = set(list_available_templates())
    if template_name not in available:
        raise HTTPException(status_code=404, detail="Template not found")
    file_path = TEMPLATE_DIR / template_name
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="Template not found")
    return FileResponse(
        path=str(file_path),
        media_type=_guess_media_type(file_path),
        filename=file_path.name,
    )


@router.get("/templates/text-preview")
def preview_template_text(template_name: str) -> dict[str, object]:
    available = set(list_available_templates())
    if template_name not in available:
        raise HTTPException(status_code=404, detail="Template not found")
    file_path = TEMPLATE_DIR / template_name
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="Template not found")
    if file_path.suffix.lower() != ".docx":
        return {"template_name": template_name, "text": "", "truncated": False}

    text = _extract_docx_text(file_path)
    max_chars = 20000
    truncated = len(text) > max_chars
    if truncated:
        text = text[:max_chars]
    return {
        "template_name": template_name,
        "text": text,
        "truncated": truncated,
    }


@router.post("/templates/match", response_model=TemplateMatchResponse)
def match_templates(request: TemplateMatchRequest) -> TemplateMatchResponse:
    templates = list_available_templates()
    matched_template, matched_by = match_template_name(
        raw_text=request.raw_text,
        file_name=request.file_name,
        templates=templates,
    )
    return TemplateMatchResponse(matched_template=matched_template, matched_by=matched_by)


@router.get("/templates/editor-schema")
def get_template_editor_schema(template_name: str) -> dict[str, object]:
    return {
        "template_name": template_name,
        "editor_schema": get_editor_schema(template_name),
    }


@router.get("/library")
def get_local_document_library(force_rebuild: bool = False) -> dict[str, object]:
    data = load_local_document_library(force_rebuild=force_rebuild)
    return {
        "library_file": str(LOCAL_DOCUMENT_LIBRARY_FILE),
        "data": data,
    }


@router.post("/library/rebuild")
def rebuild_document_library() -> dict[str, object]:
    data = rebuild_local_document_library()
    return {
        "library_file": str(LOCAL_DOCUMENT_LIBRARY_FILE),
        "data": data,
    }


@router.post("/templates/editor-prefill", response_model=EditorPrefillResponse)
def get_template_editor_prefill_data(request: EditorPrefillRequest) -> EditorPrefillResponse:
    source_file_path = _find_uploaded_file(request.source_file_id) if request.source_file_id else None
    fields = get_template_editor_prefill(
        template_name=request.template_name,
        context=request.fields.model_dump(),
        source_file_path=source_file_path,
    )
    return EditorPrefillResponse(fields=fields)


@router.post("/instrument-catalog/parse", response_model=InstrumentCatalogParseResponse)
async def parse_instrument_catalog(file: UploadFile = File(...)) -> InstrumentCatalogParseResponse:
    suffix = Path(file.filename or "").suffix.lower()
    raw_bytes = await file.read()
    if not raw_bytes:
        return InstrumentCatalogParseResponse(rows=[], names=[], total=0)

    detected_format = _detect_catalog_binary_format(raw_bytes)
    effective_format = detected_format or suffix

    if effective_format == ".xlsx":
        rows = _parse_catalog_xlsx(raw_bytes)
    elif effective_format == ".csv":
        rows = _parse_catalog_csv(raw_bytes)
    elif effective_format == ".txt":
        rows = _parse_catalog_text(raw_bytes)
    elif effective_format == ".docx":
        rows = _parse_catalog_docx(raw_bytes)
    else:
        raise HTTPException(status_code=400, detail="器具名单仅支持 .xlsx/.csv/.txt/.doc/.docx（其中 .doc 仅在可按 Word OpenXML 读取时支持）")

    names = _catalog_names_from_rows(rows)
    return InstrumentCatalogParseResponse(rows=rows, names=names, total=len(names))


@router.get("/instrument-catalog/auto-load")
def auto_load_instrument_catalog() -> dict[str, object]:
    if not INSTRUMENT_CATALOG_AUTO_ENABLED:
        return {"loaded": False, "rows": [], "names": [], "total": 0, "file_name": "", "file_path": ""}

    file_path = _find_auto_catalog_file()
    if not file_path:
        return {"loaded": False, "rows": [], "names": [], "total": 0, "file_name": "", "file_path": ""}

    raw_bytes = file_path.read_bytes()
    if not raw_bytes:
        return {"loaded": False, "rows": [], "names": [], "total": 0, "file_name": file_path.name, "file_path": str(file_path)}

    effective_format = _detect_catalog_binary_format(raw_bytes) or file_path.suffix.lower()
    if effective_format == ".xlsx":
        rows = _parse_catalog_xlsx(raw_bytes)
    elif effective_format == ".csv":
        rows = _parse_catalog_csv(raw_bytes)
    elif effective_format == ".txt":
        rows = _parse_catalog_text(raw_bytes)
    elif effective_format == ".docx":
        rows = _parse_catalog_docx(raw_bytes)
    else:
        return {"loaded": False, "rows": [], "names": [], "total": 0, "file_name": file_path.name, "file_path": str(file_path)}

    names = _catalog_names_from_rows(rows)
    return {
        "loaded": bool(names),
        "rows": rows,
        "names": names,
        "total": len(names),
        "file_name": file_path.name,
        "file_path": str(file_path),
    }


@router.get("/instrument-table/extract")
def extract_instrument_table(file_id: str) -> dict[str, object]:
    file_path = _find_uploaded_file(file_id)
    if not file_path:
        raise HTTPException(status_code=404, detail="File not found")
    raw_bytes = file_path.read_bytes()
    if not raw_bytes:
        return {"rows": [], "tsv": "", "total": 0}

    suffix = file_path.suffix.lower()
    effective_format = _detect_catalog_binary_format(raw_bytes) or suffix
    if effective_format == ".xlsx":
        parsed_rows = _parse_catalog_xlsx(raw_bytes)
    elif effective_format == ".csv":
        parsed_rows = _parse_catalog_csv(raw_bytes)
    elif effective_format == ".txt":
        parsed_rows = _parse_catalog_text(raw_bytes)
    elif effective_format == ".docx":
        parsed_rows = _extract_measurement_rows_from_docx(raw_bytes) or _parse_catalog_docx(raw_bytes)
    else:
        return {"rows": [], "tsv": "", "total": 0}

    rows: list[list[str]] = []
    for item in parsed_rows:
        if not isinstance(item, dict):
            continue
        name = _normalize_catalog_value(item.get("name", ""))
        model = _normalize_catalog_value(item.get("model", ""))
        code = _normalize_catalog_value(item.get("code", ""))
        measurement_range = _normalize_catalog_value(item.get("measurement_range", ""))
        uncertainty = _normalize_catalog_value(item.get("uncertainty", ""))
        cert_no = _normalize_catalog_value(item.get("certificate_no", ""))
        valid_date = _normalize_catalog_value(item.get("valid_date", ""))
        traceability = _normalize_catalog_value(item.get("traceability_institution", ""))
        cert_and_valid = " ".join([x for x in [cert_no, valid_date] if x]).strip()
        if not any([name, model, code, measurement_range, uncertainty, cert_and_valid, traceability]):
            continue
        if not _is_measurement_table_row_candidate(name, model, code, measurement_range, uncertainty, cert_and_valid, traceability):
            continue
        rows.append([name, model, code, measurement_range, uncertainty, cert_and_valid, traceability])

    if not rows:
        return {"rows": [], "tsv": "", "total": 0}
    header = ["计量标准器具名称", "型号/规格", "编号", "测量范围", "准确度/不确定度", "证书编号/有效期", "溯源机构"]
    table = [header, *rows]
    tsv = "\n".join(["\t".join(row) for row in table])
    return {"rows": rows, "tsv": tsv, "total": len(rows)}


def _find_report_file(report_id: str):
    matches = sorted(OUTPUT_DIR.glob(f"{report_id}__*"))
    if not matches:
        matches = sorted(OUTPUT_DIR.glob(f"{report_id}.*"))
    if not matches:
        return None
    return matches[0]


def _get_report_download_name(file_path: Path) -> str:
    name = file_path.name
    if "__" in name:
        _, tail = name.split("__", 1)
        if tail:
            return tail
    return name


def _find_uploaded_file(file_id: str | None) -> Path | None:
    if not file_id:
        return None
    matches = sorted(UPLOAD_DIR.glob(f"{file_id}.*"))
    return matches[0] if matches else None


def _find_report_batch_zip(batch_id: str) -> Path | None:
    matches = sorted(OUTPUT_DIR.glob(f"{batch_id}__excel_batch.zip"))
    return matches[0] if matches else None


def _find_auto_catalog_file() -> Path | None:
    if not INSTRUMENT_CATALOG_AUTO_DIR.exists():
        return None
    allowed_suffix = {".xlsx", ".csv", ".txt", ".doc", ".docx"}
    keywords = [str(x or "").strip().lower() for x in INSTRUMENT_CATALOG_AUTO_KEYWORDS if str(x or "").strip()]
    candidates: list[Path] = []
    for path in INSTRUMENT_CATALOG_AUTO_DIR.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() not in allowed_suffix:
            continue
        if keywords:
            name = path.name.lower()
            if not any(keyword in name for keyword in keywords):
                continue
        candidates.append(path)
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _guess_media_type(file_path):
    suffix = file_path.suffix.lower()
    if suffix == ".docx":
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    if suffix == ".html":
        return "text/html"
    if suffix == ".zip":
        return "application/zip"
    return "application/octet-stream"


_INSTRUMENT_NAME_TOKENS = ("器具名称", "设备名称", "仪器名称", "instrument name", "device name")
_INSTRUMENT_NAME_TOKEN_SET = {re.sub(r"[\s:：/\\\-_.|*（）()]+", "", value.lower()) for value in _INSTRUMENT_NAME_TOKENS}
_PLACEHOLDER_VALUES = {"", "-", "--", "—", "/", "／"}
_MODEL_TOKEN_SET = {
    "型号规格",
    "型号/规格",
    "型号",
    "规格型号",
    "modelspecification",
    "model/specification",
}
_CODE_TOKEN_SET = {
    "编号",
    "器具编号",
    "设备编号",
    "仪器编号",
    "出厂编号",
    "number",
    "serialnumber",
    "instrumentserialnumber",
}
_RANGE_TOKEN_SET = {
    "测量范围",
    "量程",
    "measurementrange",
    "range",
}
_UNCERTAINTY_TOKEN_SET = {
    "准确度等级或最大允许误差或不确定度",
    "最大允许误差",
    "不确定度",
    "uncertainty",
    "maximumpermissibleerrors",
}
_CERT_TOKEN_SET = {
    "证书编号",
    "证书号",
    "certificatenumber",
    "certificateid",
}
_VALID_DATE_TOKEN_SET = {
    "有效期限",
    "有效期",
    "validdate",
    "validuntil",
    "validity",
}
_TRACE_TOKEN_SET = {
    "溯源机构名称",
    "溯源机构",
    "traceabilityinstitution",
    "nameoftraceabilityinstitution",
}
_CATALOG_KEYS = (
    "name",
    "model",
    "code",
    "measurement_range",
    "uncertainty",
    "certificate_no",
    "valid_date",
    "traceability_institution",
)


def _parse_catalog_xlsx(raw_bytes: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="缺少 openpyxl，无法解析 Excel 清单") from exc

    try:
        wb = load_workbook(filename=io.BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel 清单解析失败：{str(exc)}") from exc

    candidates: list[dict[str, str]] = []
    for ws in wb.worksheets:
        rows = [[_normalize_catalog_value(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
        header_row_idx = -1
        column_map: dict[str, int] = {}

        for ri, row in enumerate(rows):
            candidate_map = _detect_catalog_column_map(row)
            if candidate_map.get("name", -1) >= 0:
                header_row_idx = ri
                column_map = candidate_map
                break

        if column_map.get("name", -1) >= 0:
            for row in rows[header_row_idx + 1 :]:
                candidates.append(_row_to_catalog_item(row, column_map))
            continue

        for row in rows:
            first = _pick_first_catalog_cell(row)
            if first:
                candidates.append({"name": first})

    return _finalize_catalog_rows(candidates)


def _parse_catalog_csv(raw_bytes: bytes) -> list[dict[str, str]]:
    text = _decode_catalog_bytes(raw_bytes)
    if not text:
        return []

    candidates: list[dict[str, str]] = []
    reader = csv.reader(io.StringIO(text))
    header_checked = False
    column_map: dict[str, int] = {}
    for row in reader:
        cells = [_normalize_catalog_value(cell) for cell in row]
        if not any(cells):
            continue
        if not header_checked:
            column_map = _detect_catalog_column_map(cells)
            header_checked = True
            if column_map.get("name", -1) >= 0:
                continue
        if column_map.get("name", -1) >= 0:
            candidates.append(_row_to_catalog_item(cells, column_map))
        else:
            first = _pick_first_catalog_cell(cells)
            if first:
                candidates.append({"name": first})
    return _finalize_catalog_rows(candidates)


def _parse_catalog_text(raw_bytes: bytes) -> list[dict[str, str]]:
    text = _decode_catalog_bytes(raw_bytes)
    if not text:
        return []
    candidates: list[dict[str, str]] = []
    for raw_line in text.splitlines():
        line = _normalize_catalog_value(raw_line)
        if not line:
            continue
        line = re.sub(r"^\s*\d+\s*[.)、．]\s*", "", line)
        if "," in line:
            line = line.split(",", 1)[0]
        elif "，" in line:
            line = line.split("，", 1)[0]
        elif "|" in line:
            line = line.split("|", 1)[0]
        candidates.append({"name": line})
    return _finalize_catalog_rows(candidates)


def _parse_catalog_docx(raw_bytes: bytes) -> list[dict[str, str]]:
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes), "r") as zf:
            xml_bytes = zf.read("word/document.xml")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Word 清单解析失败：{str(exc)}") from exc

    try:
        root = ET.fromstring(xml_bytes)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Word 清单内容无效：{str(exc)}") from exc

    tables = _extract_docx_table_rows(root)

    candidates: list[dict[str, str]] = []
    has_structured_table = False
    for table_rows in tables:
        header_row_idx = -1
        column_map: dict[str, int] = {}
        for ri, row in enumerate(table_rows):
            candidate_map = _detect_catalog_column_map(row)
            if candidate_map.get("name", -1) >= 0:
                header_row_idx = ri
                column_map = candidate_map
                break
        if header_row_idx >= 0:
            for row in table_rows[header_row_idx + 1 :]:
                candidates.append(_row_to_catalog_item(row, column_map))
            has_structured_table = True
            continue
        for row in table_rows:
            first = _pick_first_catalog_cell(row)
            if first:
                candidates.append({"name": first})

    # 仅在未命中结构化表格时，才走正文段落回退识别
    if not has_structured_table:
        for line in _extract_catalog_lines_from_docx_paragraphs(root):
            candidates.append({"name": line})

    return _finalize_catalog_rows(candidates)


def _decode_catalog_bytes(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return ""


def _detect_catalog_binary_format(raw_bytes: bytes) -> str:
    if len(raw_bytes) < 4:
        return ""
    if raw_bytes[:2] == b"PK":
        try:
            with zipfile.ZipFile(io.BytesIO(raw_bytes), "r") as zf:
                names = set(zf.namelist())
                if "word/document.xml" in names:
                    return ".docx"
                if "xl/workbook.xml" in names:
                    return ".xlsx"
        except Exception:
            return ""
    return ""


def _extract_catalog_lines_from_docx_paragraphs(root: ET.Element) -> list[str]:
    lines: list[str] = []
    for paragraph in root.findall(".//{*}p"):
        text = "".join([(node.text or "") for node in paragraph.findall(".//{*}t")])
        line = _normalize_catalog_value(text)
        if not line:
            continue
        # 去掉常见编号前缀（1. / 1、 / （1）/ 一、）
        line = re.sub(r"^\s*(?:\(?\d+\)?|[一二三四五六七八九十]+)\s*[.)、．）]\s*", "", line)
        line = _normalize_catalog_value(line)
        if not line:
            continue
        # 行内可能是“名称,型号,编号”或“名称|型号|编号”，正文模式先取首段作为名称候选
        if "," in line:
            line = line.split(",", 1)[0]
        elif "，" in line:
            line = line.split("，", 1)[0]
        elif "|" in line:
            line = line.split("|", 1)[0]
        line = _normalize_catalog_value(line)
        if line:
            lines.append(line)
    return lines


def _extract_docx_table_rows(root: ET.Element, preserve_paragraphs: bool = False) -> list[list[list[str]]]:
    tables: list[list[list[str]]] = []
    for tbl in root.findall(".//{*}tbl"):
        active_vmerge: dict[int, str] = {}
        rows: list[list[str]] = []
        for tr in tbl.findall("./{*}tr"):
            row: list[str] = []
            occupied: set[int] = set()
            col_idx = _get_docx_row_grid_before(tr)
            for tc in tr.findall("./{*}tc"):
                span = _get_docx_cell_grid_span(tc)
                text = _extract_docx_cell_text(tc, preserve_paragraphs=preserve_paragraphs)
                vmerge = _get_docx_cell_vmerge(tc)

                while col_idx in occupied:
                    col_idx += 1
                if col_idx < 0:
                    col_idx = 0
                if col_idx + span > len(row):
                    row.extend([""] * (col_idx + span - len(row)))

                if vmerge == "continue" and not text:
                    text = active_vmerge.get(col_idx, "")

                row[col_idx] = text
                for offset in range(span):
                    pos = col_idx + offset
                    occupied.add(pos)
                    if vmerge == "restart":
                        active_vmerge[pos] = text
                    elif vmerge != "continue":
                        active_vmerge.pop(pos, None)
                col_idx += span

            if row:
                rows.append(row)

        if rows:
            max_width = max(len(row) for row in rows)
            for row in rows:
                if len(row) < max_width:
                    row.extend([""] * (max_width - len(row)))
            tables.append(rows)

    return tables


def _extract_docx_cell_text(tc: ET.Element, preserve_paragraphs: bool = False) -> str:
    if preserve_paragraphs:
        paragraphs: list[str] = []
        for p in tc.findall("./{*}p"):
            chunks = [(node.text or "") for node in p.findall(".//{*}t")]
            line = "".join(chunks).strip()
            if line:
                paragraphs.append(line)
        return "\n".join(paragraphs).strip()
    text = "".join([(node.text or "") for node in tc.findall(".//{*}t")])
    return _normalize_catalog_value(text)


def _get_docx_cell_grid_span(tc: ET.Element) -> int:
    grid_span = tc.find("./{*}tcPr/{*}gridSpan")
    if grid_span is None:
        return 1
    try:
        span = int(grid_span.attrib.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val", "1"))
    except Exception:
        return 1
    return span if span > 0 else 1


def _get_docx_cell_vmerge(tc: ET.Element) -> str:
    vmerge = tc.find("./{*}tcPr/{*}vMerge")
    if vmerge is None:
        return ""
    return vmerge.attrib.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val", "continue") or "continue"


def _get_docx_row_grid_before(tr: ET.Element) -> int:
    grid_before = tr.find("./{*}trPr/{*}gridBefore")
    if grid_before is None:
        return 0
    try:
        value = int(grid_before.attrib.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val", "0"))
    except Exception:
        return 0
    return value if value > 0 else 0


def _pick_catalog_cell(row: list[str], col_idx: int) -> str:
    if col_idx < 0 or col_idx >= len(row):
        return ""
    return _normalize_catalog_value(row[col_idx])


def _pick_first_catalog_cell(row: list[str]) -> str:
    for cell in row:
        text = _normalize_catalog_value(cell)
        if text:
            return text
    return ""


def _normalize_catalog_value(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if text in _PLACEHOLDER_VALUES:
        return ""
    return text


def _normalize_catalog_token(value: str) -> str:
    return re.sub(r"[\s:：/\\\-_.|*（）()]+", "", str(value or "").lower())


def _token_matches_alias(token: str, alias_set: set[str]) -> bool:
    if not token:
        return False
    for alias in alias_set:
        a = _normalize_catalog_token(alias)
        if not a:
            continue
        if token == a or a in token:
            return True
    return False


def _detect_catalog_column_map(header_row: list[str]) -> dict[str, int]:
    result: dict[str, int] = {key: -1 for key in _CATALOG_KEYS}
    for idx, raw in enumerate(header_row):
        token = _normalize_catalog_token(raw)
        if not token:
            continue
        if result["name"] < 0 and _token_matches_alias(token, _INSTRUMENT_NAME_TOKEN_SET):
            result["name"] = idx
        if result["model"] < 0 and _token_matches_alias(token, _MODEL_TOKEN_SET):
            result["model"] = idx
        if result["measurement_range"] < 0 and _token_matches_alias(token, _RANGE_TOKEN_SET):
            result["measurement_range"] = idx
        if result["uncertainty"] < 0 and _token_matches_alias(token, _UNCERTAINTY_TOKEN_SET):
            result["uncertainty"] = idx
        if result["certificate_no"] < 0 and _token_matches_alias(token, _CERT_TOKEN_SET):
            result["certificate_no"] = idx
        if result["valid_date"] < 0 and _token_matches_alias(token, _VALID_DATE_TOKEN_SET):
            result["valid_date"] = idx
        if result["traceability_institution"] < 0 and _token_matches_alias(token, _TRACE_TOKEN_SET):
            result["traceability_institution"] = idx
        if (
            result["code"] < 0
            and _token_matches_alias(token, _CODE_TOKEN_SET)
            and not _token_matches_alias(token, _CERT_TOKEN_SET)
        ):
            result["code"] = idx
        if "证书编号" in token and "有效期" in token:
            if result["certificate_no"] < 0:
                result["certificate_no"] = idx
            if result["valid_date"] < 0:
                result["valid_date"] = idx
        if "certificatenumber" in token and "valid" in token:
            if result["certificate_no"] < 0:
                result["certificate_no"] = idx
            if result["valid_date"] < 0:
                result["valid_date"] = idx
    return result


def _row_to_catalog_item(row: list[str], column_map: dict[str, int]) -> dict[str, str]:
    item: dict[str, str] = {key: "" for key in _CATALOG_KEYS}
    certificate_no_idx, valid_date_idx = _resolve_certificate_and_valid_indices(row, column_map)
    for key in _CATALOG_KEYS:
        idx = int(column_map.get(key, -1))
        if key == "certificate_no":
            idx = certificate_no_idx
        elif key == "valid_date":
            idx = valid_date_idx
        if idx >= 0:
            item[key] = _pick_catalog_cell(row, idx)
    cert_no, valid_date = _split_certificate_and_valid_date(item["certificate_no"], item["valid_date"])
    item["certificate_no"] = cert_no
    item["valid_date"] = valid_date
    return item


def _extract_date_text(value: str) -> str:
    text = _normalize_catalog_value(value)
    if not text:
        return ""
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if not match:
        return ""
    return f"{match.group(1)}年{match.group(2).zfill(2)}月{match.group(3).zfill(2)}日"


def _split_certificate_and_valid_date(certificate_no: str, valid_date: str) -> tuple[str, str]:
    cert_text = _normalize_catalog_value(certificate_no)
    valid_text = _normalize_catalog_value(valid_date)
    if not cert_text and not valid_text:
        return "", ""
    if valid_text:
        return cert_text, _extract_date_text(valid_text) or valid_text
    extracted_valid = _extract_date_text(cert_text)
    if not extracted_valid:
        return cert_text, ""
    cert_only = re.sub(r"\d{4}\D+\d{1,2}\D+\d{1,2}(?:\D*日)?", "", cert_text).strip()
    cert_only = _normalize_catalog_value(cert_only)
    return cert_only, extracted_valid


def _resolve_certificate_and_valid_indices(row: list[str], column_map: dict[str, int]) -> tuple[int, int]:
    certificate_no_idx = int(column_map.get("certificate_no", -1))
    valid_date_idx = int(column_map.get("valid_date", -1))
    if certificate_no_idx >= 0 and valid_date_idx == certificate_no_idx:
        next_idx = certificate_no_idx + 1
        if next_idx < len(row):
            next_value = _normalize_catalog_value(row[next_idx])
            if next_value and _extract_date_text(next_value):
                valid_date_idx = next_idx
    return certificate_no_idx, valid_date_idx


def _extract_measurement_rows_from_docx(raw_bytes: bytes) -> list[dict[str, str]]:
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes), "r") as zf:
            xml_bytes = zf.read("word/document.xml")
    except Exception:
        return []
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return []

    tables = _extract_docx_table_rows(root, preserve_paragraphs=True)

    candidates: list[dict[str, str]] = []
    for table_rows in tables:
        header_row_idx = -1
        column_map: dict[str, int] = {}
        for ri, row in enumerate(table_rows):
            candidate_map = _detect_catalog_column_map([_normalize_catalog_value(cell) for cell in row])
            if candidate_map.get("name", -1) >= 0 and candidate_map.get("measurement_range", -1) >= 0:
                header_row_idx = ri
                column_map = candidate_map
                break
        if header_row_idx < 0:
            continue

        for row in table_rows[header_row_idx + 1 :]:
            raw_item = {key: "" for key in _CATALOG_KEYS}
            certificate_no_idx, valid_date_idx = _resolve_certificate_and_valid_indices(row, column_map)
            for key in _CATALOG_KEYS:
                idx = int(column_map.get(key, -1))
                if key == "certificate_no":
                    idx = certificate_no_idx
                elif key == "valid_date":
                    idx = valid_date_idx
                if idx < 0 or idx >= len(row):
                    continue
                raw_item[key] = str(row[idx] or "").strip()
            split_items = _split_measurement_stacked_item(raw_item)
            for item in split_items:
                candidates.append(item)

    if not candidates:
        return []
    return _finalize_catalog_rows(candidates)


def _split_measurement_stacked_item(item: dict[str, str]) -> list[dict[str, str]]:
    safe_item = {key: str((item or {}).get(key, "") or "").strip() for key in _CATALOG_KEYS}

    def split_parts(value: str) -> list[str]:
        text = str(value or "").replace("\r", "\n")
        parts = [x.strip() for x in text.split("\n") if x and x.strip()]
        return parts or [""]

    fields = {
        key: split_parts(safe_item.get(key, ""))
        for key in _CATALOG_KEYS
    }
    multi_lengths = [len(v) for v in fields.values() if len(v) > 1]
    max_len = max(multi_lengths) if multi_lengths else 1
    if max_len <= 1:
        cert_no, valid_date = _split_certificate_and_valid_date(safe_item.get("certificate_no", ""), safe_item.get("valid_date", ""))
        safe_item["certificate_no"] = cert_no
        safe_item["valid_date"] = valid_date
        return [safe_item]

    anchor_keys = ("name", "model", "code", "measurement_range", "certificate_no", "traceability_institution")
    anchor_multi = sum(1 for key in anchor_keys if len(fields.get(key, [""])) == max_len)
    if anchor_multi < 2:
        cert_no, valid_date = _split_certificate_and_valid_date(safe_item.get("certificate_no", ""), safe_item.get("valid_date", ""))
        safe_item["certificate_no"] = cert_no
        safe_item["valid_date"] = valid_date
        return [safe_item]

    rows: list[dict[str, str]] = []
    for i in range(max_len):
        row = {}
        for key in _CATALOG_KEYS:
            parts = fields.get(key, [""])
            if len(parts) == max_len:
                value = parts[i]
            elif len(parts) == 1:
                value = parts[0]
            else:
                value = parts[i] if i < len(parts) else ""
            row[key] = _normalize_catalog_value(value)
        cert_no, valid_date = _split_certificate_and_valid_date(row.get("certificate_no", ""), row.get("valid_date", ""))
        row["certificate_no"] = cert_no
        row["valid_date"] = valid_date
        rows.append(row)
    return rows


def _is_measurement_table_row_candidate(
    name: str,
    model: str,
    code: str,
    measurement_range: str,
    uncertainty: str,
    cert_and_valid: str,
    traceability: str,
) -> bool:
    name_text = _normalize_catalog_value(name)
    if not name_text:
        return False
    name_token = _normalize_catalog_token(name_text)
    if not name_token:
        return False
    if name_token in _INSTRUMENT_NAME_TOKEN_SET:
        return False
    if "mainmeasurementstandardinstrumentsusedinthiscalibration" in name_token:
        return False
    if "measurementrange" in name_token:
        return False
    if "certificatenumber" in name_token:
        return False

    model_text = _normalize_catalog_value(model)
    code_text = _normalize_catalog_value(code)
    range_text = _normalize_catalog_value(measurement_range)
    uncertainty_text = _normalize_catalog_value(uncertainty)
    cert_text = _normalize_catalog_value(cert_and_valid)
    traceability_text = _normalize_catalog_value(traceability)

    rich_fields = sum(
        1
        for value in [model_text, code_text, range_text, uncertainty_text, cert_text, traceability_text]
        if value
    )
    if rich_fields >= 2:
        return True

    has_range_like = bool(re.search(r"(?:~|～|\(|\)|mm|cm|m|℃|°c|kv|mv|v|a|μa|hz)", range_text, flags=re.IGNORECASE))
    has_cert_like = bool(re.search(r"(?:\d{4}年\d{1,2}月\d{1,2}日|[A-Za-z]\d{5,}|[A-Za-z]{1,6}[-/][A-Za-z0-9-]{3,})", cert_text))
    has_code_like = bool(re.search(r"^[A-Za-z]{1,4}[A-Za-z0-9-]{2,}$", code_text.replace(" ", "")))
    return has_range_like or has_cert_like or has_code_like


def _finalize_catalog_rows(candidates: list[dict[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for value in candidates:
        if not isinstance(value, dict):
            continue
        item = {key: _normalize_catalog_value(value.get(key, "")) for key in _CATALOG_KEYS}
        name = item.get("name", "")
        token = _normalize_catalog_token(name)
        if not name or not token or token in _INSTRUMENT_NAME_TOKEN_SET or token in seen:
            continue
        if _normalize_catalog_value(name) in _PLACEHOLDER_VALUES:
            continue
        if name in {"以上计量标准器具", "其它校准信息"}:
            continue
        if _is_catalog_header_like_item(item):
            continue
        seen.add(token)
        rows.append(item)
    return rows[:2000]


def _is_catalog_header_like_item(item: dict[str, str]) -> bool:
    name = _normalize_catalog_value((item or {}).get("name", ""))
    token = _normalize_catalog_token(name)
    if not token:
        return True

    header_tokens = {
        _normalize_catalog_token("器具总目录"),
        _normalize_catalog_token("器具名称"),
        _normalize_catalog_token("型号/规格"),
        _normalize_catalog_token("编号"),
        _normalize_catalog_token("测量范围"),
        _normalize_catalog_token("准确度等级或最大允差或不确定度"),
        _normalize_catalog_token("证书编号/有效期限"),
        _normalize_catalog_token("溯源机构名称"),
        _normalize_catalog_token("instrument name"),
        _normalize_catalog_token("model/specification"),
        _normalize_catalog_token("number"),
        _normalize_catalog_token("measurement range"),
        _normalize_catalog_token("certificate number/valid date"),
        _normalize_catalog_token("name of traceability institution"),
    }
    if token in header_tokens:
        return True
    if any(h and h in token for h in header_tokens):
        others = [
            _normalize_catalog_value((item or {}).get("model", "")),
            _normalize_catalog_value((item or {}).get("code", "")),
            _normalize_catalog_value((item or {}).get("measurement_range", "")),
            _normalize_catalog_value((item or {}).get("uncertainty", "")),
            _normalize_catalog_value((item or {}).get("certificate_no", "")),
            _normalize_catalog_value((item or {}).get("valid_date", "")),
            _normalize_catalog_value((item or {}).get("traceability_institution", "")),
        ]
        if sum(1 for x in others if x) <= 1:
            return True
    return False


def _catalog_names_from_rows(rows: list[dict[str, str]]) -> list[str]:
    names: list[str] = []
    for row in rows:
        name = _normalize_catalog_value((row or {}).get("name", ""))
        if name:
            names.append(name)
    return names[:2000]


def _build_report_validation(file_path: Path) -> ReportValidation:
    if not file_path.exists():
        return ReportValidation(ok=False)

    file_size_bytes = file_path.stat().st_size
    md5, sha256 = _calc_file_hashes(file_path)
    zip_ok = None
    missing_parts: list[str] = []

    if file_path.suffix.lower() == ".docx":
        zip_ok, missing_parts = _check_docx_package(file_path)
        ok = file_size_bytes > 0 and bool(zip_ok) and not missing_parts
    else:
        ok = file_size_bytes > 0

    return ReportValidation(
        ok=ok,
        file_size_bytes=file_size_bytes,
        md5=md5,
        sha256=sha256,
        zip_ok=zip_ok,
        missing_parts=missing_parts,
    )


def _extract_docx_text(file_path: Path) -> str:
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            xml_bytes = zf.read("word/document.xml")
    except Exception:
        return ""
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return ""

    paragraphs: list[str] = []
    for p in root.findall(".//{*}p"):
        chunks: list[str] = []
        for t in p.findall(".//{*}t"):
            if t.text:
                chunks.append(t.text)
        line = "".join(chunks).strip()
        if line:
            paragraphs.append(line)
    return "\n".join(paragraphs)


def _calc_file_hashes(file_path: Path) -> tuple[str, str]:
    md5 = hashlib.md5()
    sha256 = hashlib.sha256()
    with file_path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            md5.update(chunk)
            sha256.update(chunk)
    return md5.hexdigest(), sha256.hexdigest()


def _check_docx_package(file_path: Path) -> tuple[bool, list[str]]:
    required_parts = [
        "[Content_Types].xml",
        "_rels/.rels",
        "word/document.xml",
        "docProps/core.xml",
    ]
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            bad_entry = zf.testzip()
            if bad_entry:
                return False, required_parts
            names = set(zf.namelist())
    except Exception:
        return False, required_parts

    missing_parts = [part for part in required_parts if part not in names]
    return True, missing_parts
