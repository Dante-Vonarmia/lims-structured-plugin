import csv
import io
import re
from typing import Any
import zipfile
import xml.etree.ElementTree as ET

from fastapi import HTTPException

from .docx_structure_service import _extract_docx_table_rows

_INSTRUMENT_NAME_TOKENS = ("器具名称", "设备名称", "仪器名称", "instrument name", "device name")
_INSTRUMENT_NAME_TOKEN_SET = {re.sub(r"[\s:：/\\\-_.|*（）()]+", "", value.lower()) for value in _INSTRUMENT_NAME_TOKENS}
_PLACEHOLDER_VALUES = {"", "-", "--", "—", "/", "／"}
_MODEL_TOKEN_SET = {
    "型号规格",
    "型号/规格",
    "型号",
    "规格型号",
    "modelspecification",
    "model/specification",
}
_CODE_TOKEN_SET = {
    "编号",
    "器具编号",
    "设备编号",
    "仪器编号",
    "出厂编号",
    "number",
    "serialnumber",
    "instrumentserialnumber",
}
_RANGE_TOKEN_SET = {
    "测量范围",
    "量程",
    "measurementrange",
    "range",
}
_UNCERTAINTY_TOKEN_SET = {
    "准确度等级或最大允许误差或不确定度",
    "最大允许误差",
    "不确定度",
    "uncertainty",
    "maximumpermissibleerrors",
}
_CERT_TOKEN_SET = {
    "证书编号",
    "证书号",
    "certificatenumber",
    "certificateid",
}
_VALID_DATE_TOKEN_SET = {
    "有效期限",
    "有效期",
    "validdate",
    "validuntil",
    "validity",
}
_TRACE_TOKEN_SET = {
    "溯源机构名称",
    "溯源机构",
    "traceabilityinstitution",
    "nameoftraceabilityinstitution",
}
_CATALOG_KEYS = (
    "name",
    "model",
    "code",
    "measurement_range",
    "uncertainty",
    "certificate_no",
    "valid_date",
    "traceability_institution",
)


def _parse_catalog_xlsx(raw_bytes: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="缺少 openpyxl，无法解析 Excel 清单") from exc

    try:
        wb = load_workbook(filename=io.BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel 清单解析失败：{str(exc)}") from exc

    candidates: list[dict[str, str]] = []
    for ws in wb.worksheets:
        rows = [[_normalize_catalog_value(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
        header_row_idx = -1
        column_map: dict[str, int] = {}

        for ri, row in enumerate(rows):
            candidate_map = _detect_catalog_column_map(row)
            if candidate_map.get("name", -1) >= 0:
                header_row_idx = ri
                column_map = candidate_map
                break

        if column_map.get("name", -1) >= 0:
            for row in rows[header_row_idx + 1 :]:
                candidates.append(_row_to_catalog_item(row, column_map))
            continue

        for row in rows:
            first = _pick_first_catalog_cell(row)
            if first:
                candidates.append({"name": first})

    return _finalize_catalog_rows(candidates)


def _parse_catalog_csv(raw_bytes: bytes) -> list[dict[str, str]]:
    text = _decode_catalog_bytes(raw_bytes)
    if not text:
        return []

    candidates: list[dict[str, str]] = []
    reader = csv.reader(io.StringIO(text))
    header_checked = False
    column_map: dict[str, int] = {}
    for row in reader:
        cells = [_normalize_catalog_value(cell) for cell in row]
        if not any(cells):
            continue
        if not header_checked:
            column_map = _detect_catalog_column_map(cells)
            header_checked = True
            if column_map.get("name", -1) >= 0:
                continue
        if column_map.get("name", -1) >= 0:
            candidates.append(_row_to_catalog_item(cells, column_map))
        else:
            first = _pick_first_catalog_cell(cells)
            if first:
                candidates.append({"name": first})
    return _finalize_catalog_rows(candidates)


def _parse_catalog_text(raw_bytes: bytes) -> list[dict[str, str]]:
    text = _decode_catalog_bytes(raw_bytes)
    if not text:
        return []
    candidates: list[dict[str, str]] = []
    for raw_line in text.splitlines():
        line = _normalize_catalog_value(raw_line)
        if not line:
            continue
        line = re.sub(r"^\s*\d+\s*[.)、．]\s*", "", line)
        if "," in line:
            line = line.split(",", 1)[0]
        elif "，" in line:
            line = line.split("，", 1)[0]
        elif "|" in line:
            line = line.split("|", 1)[0]
        candidates.append({"name": line})
    return _finalize_catalog_rows(candidates)


def _parse_catalog_docx(raw_bytes: bytes) -> list[dict[str, str]]:
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes), "r") as zf:
            xml_bytes = zf.read("word/document.xml")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Word 清单解析失败：{str(exc)}") from exc

    try:
        root = ET.fromstring(xml_bytes)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Word 清单内容无效：{str(exc)}") from exc

    tables = _extract_docx_table_rows(root)

    candidates: list[dict[str, str]] = []
    has_structured_table = False
    for table_rows in tables:
        header_row_idx = -1
        column_map: dict[str, int] = {}
        for ri, row in enumerate(table_rows):
            candidate_map = _detect_catalog_column_map(row)
            if candidate_map.get("name", -1) >= 0:
                header_row_idx = ri
                column_map = candidate_map
                break
        if header_row_idx >= 0:
            for row in table_rows[header_row_idx + 1 :]:
                candidates.append(_row_to_catalog_item(row, column_map))
            has_structured_table = True
            continue
        for row in table_rows:
            first = _pick_first_catalog_cell(row)
            if first:
                candidates.append({"name": first})

    # 仅在未命中结构化表格时，才走正文段落回退识别
    if not has_structured_table:
        for line in _extract_catalog_lines_from_docx_paragraphs(root):
            candidates.append({"name": line})

    return _finalize_catalog_rows(candidates)


def _decode_catalog_bytes(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return ""


def _detect_catalog_binary_format(raw_bytes: bytes) -> str:
    if len(raw_bytes) < 4:
        return ""
    if raw_bytes[:2] == b"PK":
        try:
            with zipfile.ZipFile(io.BytesIO(raw_bytes), "r") as zf:
                names = set(zf.namelist())
                if "word/document.xml" in names:
                    return ".docx"
                if "xl/workbook.xml" in names:
                    return ".xlsx"
        except Exception:
            return ""
    return ""


def _extract_catalog_lines_from_docx_paragraphs(root: ET.Element) -> list[str]:
    lines: list[str] = []
    for paragraph in root.findall(".//{*}p"):
        text = "".join([(node.text or "") for node in paragraph.findall(".//{*}t")])
        line = _normalize_catalog_value(text)
        if not line:
            continue
        # 去掉常见编号前缀（1. / 1、 / （1）/ 一、）
        line = re.sub(r"^\s*(?:\(?\d+\)?|[一二三四五六七八九十]+)\s*[.)、．）]\s*", "", line)
        line = _normalize_catalog_value(line)
        if not line:
            continue
        # 行内可能是“名称,型号,编号”或“名称|型号|编号”，正文模式先取首段作为名称候选
        if "," in line:
            line = line.split(",", 1)[0]
        elif "，" in line:
            line = line.split("，", 1)[0]
        elif "|" in line:
            line = line.split("|", 1)[0]
        line = _normalize_catalog_value(line)
        if line:
            lines.append(line)
    return lines


def _pick_catalog_cell(row: list[str], col_idx: int) -> str:
    if col_idx < 0 or col_idx >= len(row):
        return ""
    return _normalize_catalog_value(row[col_idx])


def _pick_first_catalog_cell(row: list[str]) -> str:
    for cell in row:
        text = _normalize_catalog_value(cell)
        if text:
            return text
    return ""


def _normalize_catalog_value(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if text in _PLACEHOLDER_VALUES:
        return ""
    return text


def _normalize_catalog_token(value: str) -> str:
    return re.sub(r"[\s:：/\\\-_.|*（）()]+", "", str(value or "").lower())


def _token_matches_alias(token: str, alias_set: set[str]) -> bool:
    if not token:
        return False
    for alias in alias_set:
        a = _normalize_catalog_token(alias)
        if not a:
            continue
        if token == a or a in token:
            return True
    return False


def _detect_catalog_column_map(header_row: list[str]) -> dict[str, int]:
    result: dict[str, int] = {key: -1 for key in _CATALOG_KEYS}
    for idx, raw in enumerate(header_row):
        token = _normalize_catalog_token(raw)
        if not token:
            continue
        if result["name"] < 0 and _token_matches_alias(token, _INSTRUMENT_NAME_TOKEN_SET):
            result["name"] = idx
        if result["model"] < 0 and _token_matches_alias(token, _MODEL_TOKEN_SET):
            result["model"] = idx
        if result["measurement_range"] < 0 and _token_matches_alias(token, _RANGE_TOKEN_SET):
            result["measurement_range"] = idx
        if result["uncertainty"] < 0 and _token_matches_alias(token, _UNCERTAINTY_TOKEN_SET):
            result["uncertainty"] = idx
        if result["certificate_no"] < 0 and _token_matches_alias(token, _CERT_TOKEN_SET):
            result["certificate_no"] = idx
        if result["valid_date"] < 0 and _token_matches_alias(token, _VALID_DATE_TOKEN_SET):
            result["valid_date"] = idx
        if result["traceability_institution"] < 0 and _token_matches_alias(token, _TRACE_TOKEN_SET):
            result["traceability_institution"] = idx
        if (
            result["code"] < 0
            and _token_matches_alias(token, _CODE_TOKEN_SET)
            and not _token_matches_alias(token, _CERT_TOKEN_SET)
        ):
            result["code"] = idx
        if "证书编号" in token and "有效期" in token:
            if result["certificate_no"] < 0:
                result["certificate_no"] = idx
            if result["valid_date"] < 0:
                result["valid_date"] = idx
        if "certificatenumber" in token and "valid" in token:
            if result["certificate_no"] < 0:
                result["certificate_no"] = idx
            if result["valid_date"] < 0:
                result["valid_date"] = idx
    return result


def _row_to_catalog_item(row: list[str], column_map: dict[str, int]) -> dict[str, str]:
    item: dict[str, str] = {key: "" for key in _CATALOG_KEYS}
    certificate_no_idx, valid_date_idx = _resolve_certificate_and_valid_indices(row, column_map)
    for key in _CATALOG_KEYS:
        idx = int(column_map.get(key, -1))
        if key == "certificate_no":
            idx = certificate_no_idx
        elif key == "valid_date":
            idx = valid_date_idx
        if idx >= 0:
            item[key] = _pick_catalog_cell(row, idx)
    cert_no, valid_date = _split_certificate_and_valid_date(item["certificate_no"], item["valid_date"])
    item["certificate_no"] = cert_no
    item["valid_date"] = valid_date
    return item


def _extract_date_text(value: str) -> str:
    text = _normalize_catalog_value(value)
    if not text:
        return ""
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if not match:
        return ""
    return f"{match.group(1)}年{match.group(2).zfill(2)}月{match.group(3).zfill(2)}日"


def _split_certificate_and_valid_date(certificate_no: str, valid_date: str) -> tuple[str, str]:
    cert_text = _normalize_catalog_value(certificate_no)
    valid_text = _normalize_catalog_value(valid_date)
    if not cert_text and not valid_text:
        return "", ""
    if valid_text:
        return cert_text, _extract_date_text(valid_text) or valid_text
    extracted_valid = _extract_date_text(cert_text)
    if not extracted_valid:
        return cert_text, ""
    cert_only = re.sub(r"\d{4}\D+\d{1,2}\D+\d{1,2}(?:\D*日)?", "", cert_text).strip()
    cert_only = _normalize_catalog_value(cert_only)
    return cert_only, extracted_valid


def _resolve_certificate_and_valid_indices(row: list[str], column_map: dict[str, int]) -> tuple[int, int]:
    certificate_no_idx = int(column_map.get("certificate_no", -1))
    valid_date_idx = int(column_map.get("valid_date", -1))
    if certificate_no_idx >= 0 and valid_date_idx == certificate_no_idx:
        next_idx = certificate_no_idx + 1
        if next_idx < len(row):
            next_value = _normalize_catalog_value(row[next_idx])
            if next_value and _extract_date_text(next_value):
                valid_date_idx = next_idx
    return certificate_no_idx, valid_date_idx


def _extract_measurement_rows_from_docx(raw_bytes: bytes) -> list[dict[str, str]]:
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes), "r") as zf:
            xml_bytes = zf.read("word/document.xml")
    except Exception:
        return []
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return []

    tables = _extract_docx_table_rows(root, preserve_paragraphs=True)

    candidates: list[dict[str, str]] = []
    for table_rows in tables:
        header_row_idx = -1
        column_map: dict[str, int] = {}
        for ri, row in enumerate(table_rows):
            candidate_map = _detect_catalog_column_map([_normalize_catalog_value(cell) for cell in row])
            if candidate_map.get("name", -1) >= 0 and candidate_map.get("measurement_range", -1) >= 0:
                header_row_idx = ri
                column_map = candidate_map
                break
        if header_row_idx < 0:
            continue

        for row in table_rows[header_row_idx + 1 :]:
            raw_item = {key: "" for key in _CATALOG_KEYS}
            certificate_no_idx, valid_date_idx = _resolve_certificate_and_valid_indices(row, column_map)
            for key in _CATALOG_KEYS:
                idx = int(column_map.get(key, -1))
                if key == "certificate_no":
                    idx = certificate_no_idx
                elif key == "valid_date":
                    idx = valid_date_idx
                if idx < 0 or idx >= len(row):
                    continue
                raw_item[key] = str(row[idx] or "").strip()
            split_items = _split_measurement_stacked_item(raw_item)
            for item in split_items:
                candidates.append(item)

    if not candidates:
        return []
    return _finalize_catalog_rows(candidates)


def _split_measurement_stacked_item(item: dict[str, str]) -> list[dict[str, str]]:
    safe_item = {key: str((item or {}).get(key, "") or "").strip() for key in _CATALOG_KEYS}

    def split_parts(value: str) -> list[str]:
        text = str(value or "").replace("\r", "\n")
        parts = [x.strip() for x in text.split("\n") if x and x.strip()]
        return parts or [""]

    fields = {
        key: split_parts(safe_item.get(key, ""))
        for key in _CATALOG_KEYS
    }
    multi_lengths = [len(v) for v in fields.values() if len(v) > 1]
    max_len = max(multi_lengths) if multi_lengths else 1
    if max_len <= 1:
        cert_no, valid_date = _split_certificate_and_valid_date(safe_item.get("certificate_no", ""), safe_item.get("valid_date", ""))
        safe_item["certificate_no"] = cert_no
        safe_item["valid_date"] = valid_date
        return [safe_item]

    anchor_keys = ("name", "model", "code", "measurement_range", "certificate_no", "traceability_institution")
    anchor_multi = sum(1 for key in anchor_keys if len(fields.get(key, [""])) == max_len)
    if anchor_multi < 2:
        cert_no, valid_date = _split_certificate_and_valid_date(safe_item.get("certificate_no", ""), safe_item.get("valid_date", ""))
        safe_item["certificate_no"] = cert_no
        safe_item["valid_date"] = valid_date
        return [safe_item]

    rows: list[dict[str, str]] = []
    for i in range(max_len):
        row = {}
        for key in _CATALOG_KEYS:
            parts = fields.get(key, [""])
            if len(parts) == max_len:
                value = parts[i]
            elif len(parts) == 1:
                value = parts[0]
            else:
                value = parts[i] if i < len(parts) else ""
            row[key] = _normalize_catalog_value(value)
        cert_no, valid_date = _split_certificate_and_valid_date(row.get("certificate_no", ""), row.get("valid_date", ""))
        row["certificate_no"] = cert_no
        row["valid_date"] = valid_date
        rows.append(row)
    return rows


def _is_measurement_table_row_candidate(
    name: str,
    model: str,
    code: str,
    measurement_range: str,
    uncertainty: str,
    cert_and_valid: str,
    traceability: str,
) -> bool:
    name_text = _normalize_catalog_value(name)
    if not name_text:
        return False
    name_token = _normalize_catalog_token(name_text)
    if not name_token:
        return False
    if name_token in _INSTRUMENT_NAME_TOKEN_SET:
        return False
    if "mainmeasurementstandardinstrumentsusedinthiscalibration" in name_token:
        return False
    if "measurementrange" in name_token:
        return False
    if "certificatenumber" in name_token:
        return False

    model_text = _normalize_catalog_value(model)
    code_text = _normalize_catalog_value(code)
    range_text = _normalize_catalog_value(measurement_range)
    uncertainty_text = _normalize_catalog_value(uncertainty)
    cert_text = _normalize_catalog_value(cert_and_valid)
    traceability_text = _normalize_catalog_value(traceability)

    rich_fields = sum(
        1
        for value in [model_text, code_text, range_text, uncertainty_text, cert_text, traceability_text]
        if value
    )
    if rich_fields >= 2:
        return True

    has_range_like = bool(re.search(r"(?:~|～|\(|\)|mm|cm|m|℃|°c|kv|mv|v|a|μa|hz)", range_text, flags=re.IGNORECASE))
    has_cert_like = bool(re.search(r"(?:\d{4}年\d{1,2}月\d{1,2}日|[A-Za-z]\d{5,}|[A-Za-z]{1,6}[-/][A-Za-z0-9-]{3,})", cert_text))
    has_code_like = bool(re.search(r"^[A-Za-z]{1,4}[A-Za-z0-9-]{2,}$", code_text.replace(" ", "")))
    return has_range_like or has_cert_like or has_code_like


def _finalize_catalog_rows(candidates: list[dict[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for value in candidates:
        if not isinstance(value, dict):
            continue
        item = {key: _normalize_catalog_value(value.get(key, "")) for key in _CATALOG_KEYS}
        name = item.get("name", "")
        token = _normalize_catalog_token(name)
        if not name or not token or token in _INSTRUMENT_NAME_TOKEN_SET or token in seen:
            continue
        if _normalize_catalog_value(name) in _PLACEHOLDER_VALUES:
            continue
        if name in {"以上计量标准器具", "其它校准信息"}:
            continue
        if _is_catalog_header_like_item(item):
            continue
        seen.add(token)
        rows.append(item)
    return rows[:2000]


def _is_catalog_header_like_item(item: dict[str, str]) -> bool:
    name = _normalize_catalog_value((item or {}).get("name", ""))
    token = _normalize_catalog_token(name)
    if not token:
        return True

    header_tokens = {
        _normalize_catalog_token("器具总目录"),
        _normalize_catalog_token("器具名称"),
        _normalize_catalog_token("型号/规格"),
        _normalize_catalog_token("编号"),
        _normalize_catalog_token("测量范围"),
        _normalize_catalog_token("准确度等级或最大允差或不确定度"),
        _normalize_catalog_token("证书编号/有效期限"),
        _normalize_catalog_token("溯源机构名称"),
        _normalize_catalog_token("instrument name"),
        _normalize_catalog_token("model/specification"),
        _normalize_catalog_token("number"),
        _normalize_catalog_token("measurement range"),
        _normalize_catalog_token("certificate number/valid date"),
        _normalize_catalog_token("name of traceability institution"),
    }
    if token in header_tokens:
        return True
    if any(h and h in token for h in header_tokens):
        others = [
            _normalize_catalog_value((item or {}).get("model", "")),
            _normalize_catalog_value((item or {}).get("code", "")),
            _normalize_catalog_value((item or {}).get("measurement_range", "")),
            _normalize_catalog_value((item or {}).get("uncertainty", "")),
            _normalize_catalog_value((item or {}).get("certificate_no", "")),
            _normalize_catalog_value((item or {}).get("valid_date", "")),
            _normalize_catalog_value((item or {}).get("traceability_institution", "")),
        ]
        if sum(1 for x in others if x) <= 1:
            return True
    return False


def _catalog_names_from_rows(rows: list[dict[str, str]]) -> list[str]:
    names: list[str] = []
    for row in rows:
        name = _normalize_catalog_value((row or {}).get("name", ""))
        if name:
            names.append(name)
    return names[:2000]
