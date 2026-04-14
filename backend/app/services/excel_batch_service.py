from pathlib import Path
import re
from typing import Any

from .local_document_library_service import load_local_document_library
from .template_service import list_available_templates, match_template_name

FIELD_ALIASES: dict[str, set[str]] = {
    "template_name": {"template_name", "模板", "模板名称", "template"},
    "device_name": {"device_name", "器具名称", "设备名称", "仪器名称", "name"},
    "device_model": {"device_model", "型号规格", "规格型号", "型号", "规格", "model"},
    "device_code": {"device_code", "器具编号", "设备编号", "资产编号", "编号", "serial", "code"},
    "manufacturer": {"manufacturer", "生产厂商", "制造厂商", "制造厂/商", "厂家", "厂商"},
    "use_department": {"use_department", "使用部门", "部门", "使用科室", "科室"},
    "certificate_no": {"certificate_no", "证书编号", "证书号", "缆专检号", "certificate"},
    "client_name": {"client_name", "委托单位", "客户名称", "客户"},
    "receive_date": {"receive_date", "收样日期", "收样时间"},
    "calibration_date": {"calibration_date", "校准日期", "检定日期", "检测日期"},
    "location": {"location", "地点", "校准地点", "检定地点"},
    "temperature": {"temperature", "温度"},
    "humidity": {"humidity", "湿度"},
    "source_code": {"source_code", "模板代号", "来源代号", "模板编号"},
}

EMPTY_FIELDS = {
    "device_name": "",
    "device_model": "",
    "device_code": "",
    "manufacturer": "",
    "use_department": "",
    "certificate_no": "",
    "client_name": "",
    "receive_date": "",
    "calibration_date": "",
    "location": "",
    "temperature": "",
    "humidity": "",
    "section2_u_mm": "",
    "section2_value_mm": "",
    "section3_u_g": "",
    "section3_value_g": "",
    "section4_u_g": "",
    "hammer_actual_row_1": "",
    "hammer_actual_row_2": "",
    "hammer_actual_row_3": "",
    "source_profile": "excel_row",
    "source_profile_label": "Excel行",
    "device_group_count": "1",
    "device_group_summary": "",
    "has_measurement_scope": "0",
    "raw_record": "",
}


def parse_excel_rows(
    file_path: Path,
    sheet_name: str | None = None,
    default_template_name: str | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    target_sheets, _sheet_names, _current_sheet_name = load_excel_sheets(
        file_path=file_path,
        sheet_name=sheet_name,
        runtime_label="batch mode",
    )

    templates = list_available_templates()
    excel_lookup = build_excel_field_lookup(target_sheets)
    local_lookup = build_local_field_lookup()

    items: list[dict[str, Any]] = []
    errors: list[str] = []
    for ws in target_sheets:
        rows = ws.get("rows", [])
        ws_title = str(ws.get("title", "") or "")
        header_idx = detect_header_row_index(rows)
        if header_idx < 0:
            errors.append(f"[{ws_title}] 表头为空")
            continue

        normalized_headers = [normalize_header(v) for v in rows[header_idx]]
        for physical_idx in range(header_idx + 1, len(rows)):
            logical_row = physical_idx + 1
            values = rows[physical_idx]
            if not any(values):
                continue

            row_data = row_to_data_map(normalized_headers, values)
            fields = dict(EMPTY_FIELDS)
            for field_key in fields.keys():
                value = extract_field_value(row_data, field_key)
                if value:
                    fields[field_key] = value
            enrich_fields_from_lookup(fields, excel_lookup, local_lookup)

            if is_placeholder_record_row(fields):
                continue

            fields["raw_record"] = build_raw_record(row_data)

            template_name = resolve_template_for_row(
                row_data=row_data,
                fields=fields,
                templates=templates,
                default_template_name=default_template_name,
            )

            if not template_name:
                errors.append(
                    f"[{ws_title}] 第 {logical_row} 行缺少模板信息（template_name/source_code/default_template_name）"
                )
                continue

            if template_name not in templates:
                errors.append(f"[{ws_title}] 第 {logical_row} 行模板不存在：{template_name}")
                continue

            row_name = fields.get("certificate_no") or fields.get("device_code") or fields.get("device_name") or f"row_{logical_row}"
            items.append(
                {
                    "sheet_name": ws_title,
                    "row_number": logical_row,
                    "row_name": sanitize_file_name(row_name),
                    "template_name": template_name,
                    "fields": fields,
                }
            )

    return items, errors


def inspect_excel_records(
    file_path: Path,
    sheet_name: str | None = None,
    default_template_name: str | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    target_sheets, _sheet_names, _current_sheet_name = load_excel_sheets(
        file_path=file_path,
        sheet_name=sheet_name,
        runtime_label="inspect mode",
    )

    templates = list_available_templates()
    excel_lookup = build_excel_field_lookup(target_sheets)
    local_lookup = build_local_field_lookup()

    items: list[dict[str, Any]] = []
    errors: list[str] = []
    for ws in target_sheets:
        rows = ws.get("rows", [])
        ws_title = str(ws.get("title", "") or "")
        header_idx = detect_header_row_index(rows)
        if header_idx < 0:
            errors.append(f"[{ws_title}] 表头为空")
            continue

        normalized_headers = [normalize_header(v) for v in rows[header_idx]]
        for physical_idx in range(header_idx + 1, len(rows)):
            logical_row = physical_idx + 1
            values = rows[physical_idx]
            if not any(values):
                continue

            row_data = row_to_data_map(normalized_headers, values)
            fields = dict(EMPTY_FIELDS)
            for field_key in fields.keys():
                value = extract_field_value(row_data, field_key)
                if value:
                    fields[field_key] = value
            enrich_fields_from_lookup(fields, excel_lookup, local_lookup)

            if is_placeholder_record_row(fields):
                continue

            fields["raw_record"] = build_raw_record(row_data)

            template_name = resolve_template_for_row(
                row_data=row_data,
                fields=fields,
                templates=templates,
                default_template_name=default_template_name,
            )

            row_error = ""
            if template_name and template_name not in templates:
                row_error = f"[{ws_title}] 第 {logical_row} 行模板不存在：{template_name}"
                errors.append(row_error)

            row_name = fields.get("certificate_no") or fields.get("device_code") or fields.get("device_name") or f"row_{logical_row}"
            items.append(
                {
                    "sheet_name": ws_title,
                    "row_number": logical_row,
                    "row_name": sanitize_file_name(row_name),
                    "template_name": template_name,
                    "fields": fields,
                    "error": row_error,
                }
            )

    return items, errors


def preview_excel_sheet(
    file_path: Path,
    sheet_name: str | None = None,
    max_rows: int = 80,
    max_cols: int = 20,
) -> dict[str, Any]:
    target_sheets, sheet_names, current_sheet_name = load_excel_sheets(
        file_path=file_path,
        sheet_name=sheet_name,
        runtime_label="preview mode",
    )
    if not target_sheets:
        return {
            "sheet_names": [],
            "sheet_name": "",
            "headers": [],
            "rows": [],
            "row_numbers": [],
            "total_rows": 0,
            "truncated": False,
        }

    ws = target_sheets[0]
    rows = ws.get("rows", [])
    non_empty_rows = [(idx, row) for idx, row in enumerate(rows) if any(cell for cell in row)]
    if not non_empty_rows:
        return {
            "sheet_names": sheet_names,
            "sheet_name": current_sheet_name,
            "headers": [],
            "rows": [],
            "row_numbers": [],
            "total_rows": 0,
            "truncated": False,
        }
    title = ""
    header_idx = detect_header_row_index(rows, max_scan=60, min_non_empty=2)
    if header_idx < 0:
        header_idx = non_empty_rows[0][0]
    for idx, row in non_empty_rows:
        if idx >= header_idx:
            break
        limited = row[:max_cols]
        non_empty_cells = [cell for cell in limited if cell]
        if len(non_empty_cells) == 1 and len(non_empty_cells[0]) >= 8:
            title = non_empty_cells[0]
            break

    header_row = rows[header_idx][:max_cols]
    headers = [v if v else f"列{idx + 1}" for idx, v in enumerate(header_row)]
    records: list[list[str]] = []
    row_numbers: list[int] = []
    total_rows = 0
    truncated = False
    for idx in range(header_idx + 1, len(rows)):
        values = rows[idx][:max_cols]
        if not any(values):
            continue
        total_rows += 1
        if len(records) >= max_rows:
            truncated = True
            continue
        records.append(values)
        row_numbers.append(idx + 1)

    return {
        "sheet_names": sheet_names,
        "sheet_name": current_sheet_name,
        "title": title,
        "headers": headers,
        "rows": records,
        "row_numbers": row_numbers,
        "total_rows": total_rows,
        "truncated": truncated,
    }


def detect_header_row_index(rows: list[list[str]], max_scan: int = 60, min_non_empty: int = 2) -> int:
    non_empty_rows = [(idx, row) for idx, row in enumerate(rows) if any(cell for cell in row)]
    if not non_empty_rows:
        return -1
    scanned = non_empty_rows[:max_scan]

    # First pass: match header-like rows by known aliases/keywords.
    best_header_idx = scanned[0][0]
    best_header_score = -10**9
    best_exact_hits = -1
    best_non_empty = -1
    for idx, row in scanned:
        score, exact_hits, non_empty_count = score_header_row(row)
        if (
            score > best_header_score
            or (score == best_header_score and exact_hits > best_exact_hits)
            or (score == best_header_score and exact_hits == best_exact_hits and non_empty_count > best_non_empty)
        ):
            best_header_score = score
            best_exact_hits = exact_hits
            best_non_empty = non_empty_count
            best_header_idx = idx

    # Exact alias hits >=2 means a highly confident header row.
    if best_exact_hits >= 2:
        return best_header_idx

    # Fallback: keep old behavior (row with most non-empty cells).
    header_idx = scanned[0][0]
    fallback_score = -1
    for idx, row in scanned:
        score = sum(1 for cell in row if cell)
        if score >= min_non_empty and score > fallback_score:
            fallback_score = score
            header_idx = idx
    return header_idx


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


HEADER_ALIAS_LABELS: set[str] = set()
for alias_set in FIELD_ALIASES.values():
    for alias in alias_set:
        normalized = normalize_header(alias)
        if normalized:
            HEADER_ALIAS_LABELS.add(normalized)
for extra_header in [
    "序号",
    "资产编号",
    "设备编号",
    "设备名称",
    "规格型号",
    "器具编号",
    "计量单位",
    "购置日期",
    "启用日期",
    "使用日期",
    "设备状态",
    "账面原值",
    "地点",
    "检验室",
    "原检验室",
]:
    normalized = normalize_header(extra_header)
    if normalized:
        HEADER_ALIAS_LABELS.add(normalized)

HEADER_KEYWORD_HINTS = (
    "编号",
    "名称",
    "型号",
    "日期",
    "模板",
    "设备",
    "器具",
    "证书",
    "厂商",
    "厂家",
    "单位",
    "地址",
    "状态",
    "地点",
    "计量",
    "检验",
)


def score_header_row(row: list[str]) -> tuple[int, int, int]:
    normalized_cells = [normalize_header(cell) for cell in row if str(cell or "").strip()]
    if not normalized_cells:
        return -10**6, 0, 0

    exact_hits = sum(1 for cell in normalized_cells if cell in HEADER_ALIAS_LABELS)
    keyword_hits = sum(1 for cell in normalized_cells if any(key in cell for key in HEADER_KEYWORD_HINTS))
    data_like_hits = sum(1 for cell in normalized_cells if re.fullmatch(r"[0-9a-z./:_-]+", cell))
    non_empty_count = len(normalized_cells)

    score = exact_hits * 100 + keyword_hits * 12 + non_empty_count - data_like_hits * 3
    return score, exact_hits, non_empty_count


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def row_to_data_map(headers: list[str], values: list[str]) -> dict[str, str]:
    data: dict[str, str] = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        value = values[idx] if idx < len(values) else ""
        if value:
            data[header] = value
    return data


def extract_field_value(row_data: dict[str, str], field_key: str) -> str:
    aliases = FIELD_ALIASES.get(field_key, set())
    for alias in aliases:
        normalized_alias = normalize_header(alias)
        value = row_data.get(normalized_alias, "")
        if value:
            return value.strip()
    return ""


def build_raw_record(row_data: dict[str, str]) -> str:
    lines: list[str] = []
    for key, value in row_data.items():
        if value:
            lines.append(f"{key}: {value}")
    return "\n".join(lines)


def resolve_template_for_row(
    row_data: dict[str, str],
    fields: dict[str, str],
    templates: list[str],
    default_template_name: str | None = None,
) -> str:
    template_name = extract_field_value(row_data, "template_name")
    if template_name:
        return template_name

    source_code = extract_field_value(row_data, "source_code")
    if source_code:
        matched_by_code = resolve_template_by_code(source_code, templates)
        if matched_by_code:
            return matched_by_code

    raw_text = fields.get("raw_record", "")
    file_name_hint = fields.get("device_name", "") or None
    matched_template, _ = match_template_name(
        raw_text=raw_text,
        file_name=file_name_hint,
        templates=templates,
    )
    if matched_template:
        return matched_template

    if default_template_name:
        return default_template_name

    blank_template = resolve_blank_template(templates)
    if blank_template:
        return blank_template
    return ""


def build_local_field_lookup() -> dict[str, dict[str, set[str]]]:
    result: dict[str, dict[str, set[str]]] = {
        "manufacturer_by_code": {},
        "manufacturer_by_name_model": {},
        "device_name_by_code": {},
        "device_model_by_code": {},
    }
    try:
        library = load_local_document_library(force_rebuild=False)
    except Exception:
        return result

    records = library.get("raw_records", []) if isinstance(library, dict) else []
    if not isinstance(records, list):
        return result

    for record in records:
        if not isinstance(record, dict):
            continue
        basic = record.get("basic_values", {})
        if not isinstance(basic, dict):
            continue
        add_lookup_entry(result, basic)

    return result


def build_excel_field_lookup(target_sheets: list[Any]) -> dict[str, dict[str, set[str]]]:
    result: dict[str, dict[str, set[str]]] = {
        "manufacturer_by_code": {},
        "manufacturer_by_name_model": {},
        "device_name_by_code": {},
        "device_model_by_code": {},
    }
    for ws in target_sheets:
        rows = ws.get("rows", []) if isinstance(ws, dict) else []
        if not isinstance(rows, list):
            continue
        header_idx = detect_header_row_index(rows)
        if header_idx < 0:
            continue
        headers = [normalize_header(v) for v in rows[header_idx]]
        for row_values in rows[header_idx + 1 :]:
            if not any(row_values):
                continue
            row_data = row_to_data_map(headers, row_values)
            fields = dict(EMPTY_FIELDS)
            for field_key in ("device_name", "device_model", "device_code", "manufacturer"):
                value = extract_field_value(row_data, field_key)
                if value:
                    fields[field_key] = value
            add_lookup_entry(result, fields)
    return result


def load_excel_sheets(
    file_path: Path,
    sheet_name: str | None = None,
    runtime_label: str = "excel mode",
) -> tuple[list[dict[str, Any]], list[str], str]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            raise RuntimeError(f"openpyxl is required for Excel {runtime_label}") from exc

        wb = load_workbook(filename=str(file_path), data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            source_sheets = [wb[sheet_name]]
        else:
            source_sheets = list(wb.worksheets)
        sheets = [
            {
                "title": ws.title,
                "rows": [[normalize_cell(v) for v in row] for row in ws.iter_rows(values_only=True)],
            }
            for ws in source_sheets
        ]
        current_sheet_name = sheets[0]["title"] if sheets else ""
        return sheets, list(wb.sheetnames), current_sheet_name

    if suffix == ".xls":
        try:
            import xlrd
        except Exception as exc:  # pragma: no cover
            raise RuntimeError(f"xlrd is required for Excel {runtime_label}") from exc

        wb = xlrd.open_workbook(filename=str(file_path))
        sheet_names = wb.sheet_names()
        if sheet_name and sheet_name in sheet_names:
            source_sheets = [wb.sheet_by_name(sheet_name)]
        else:
            source_sheets = [wb.sheet_by_index(i) for i in range(wb.nsheets)]
        sheets = [
            {
                "title": ws.name,
                "rows": [
                    [normalize_cell(ws.cell_value(row_idx, col_idx)) for col_idx in range(ws.ncols)]
                    for row_idx in range(ws.nrows)
                ],
            }
            for ws in source_sheets
        ]
        current_sheet_name = sheets[0]["title"] if sheets else ""
        return sheets, sheet_names, current_sheet_name

    raise RuntimeError(f"Unsupported excel suffix: {suffix}")


def add_lookup_entry(lookup: dict[str, dict[str, set[str]]], fields: dict[str, str]) -> None:
    code = normalize_lookup_token(fields.get("device_code", ""))
    name = normalize_lookup_token(fields.get("device_name", ""))
    model = normalize_lookup_token(fields.get("device_model", ""))
    manufacturer = normalize_cell(fields.get("manufacturer", ""))
    device_name = normalize_cell(fields.get("device_name", ""))
    device_model = normalize_cell(fields.get("device_model", ""))

    if code and manufacturer:
        lookup["manufacturer_by_code"].setdefault(code, set()).add(manufacturer)
    if name and model and manufacturer:
        pair_key = f"{name}|{model}"
        lookup["manufacturer_by_name_model"].setdefault(pair_key, set()).add(manufacturer)
    if code and device_name:
        lookup["device_name_by_code"].setdefault(code, set()).add(device_name)
    if code and device_model:
        lookup["device_model_by_code"].setdefault(code, set()).add(device_model)


def enrich_fields_from_lookup(
    fields: dict[str, str],
    primary_lookup: dict[str, dict[str, set[str]]],
    fallback_lookup: dict[str, dict[str, set[str]]] | None = None,
) -> None:
    if not fields:
        return
    code = normalize_lookup_token(fields.get("device_code", ""))
    name = normalize_lookup_token(fields.get("device_name", ""))
    model = normalize_lookup_token(fields.get("device_model", ""))
    lookups = [primary_lookup]
    if fallback_lookup:
        lookups.append(fallback_lookup)

    if not normalize_cell(fields.get("device_name", "")) and code:
        for lookup in lookups:
            matched = pick_unique_lookup(lookup.get("device_name_by_code", {}), code)
            if matched:
                fields["device_name"] = matched
                break
    if not normalize_cell(fields.get("device_model", "")) and code:
        for lookup in lookups:
            matched = pick_unique_lookup(lookup.get("device_model_by_code", {}), code)
            if matched:
                fields["device_model"] = matched
                break

    if not normalize_cell(fields.get("manufacturer", "")):
        for lookup in lookups:
            if code:
                matched = pick_unique_lookup(lookup.get("manufacturer_by_code", {}), code)
                if matched:
                    fields["manufacturer"] = matched
                    return
            if name and model:
                pair_key = f"{name}|{model}"
                matched = pick_unique_lookup(lookup.get("manufacturer_by_name_model", {}), pair_key)
                if matched:
                    fields["manufacturer"] = matched
                    return


def pick_unique_lookup(mapping: dict[str, set[str]], key: str) -> str:
    values = mapping.get(key, set())
    if len(values) != 1:
        return ""
    return next(iter(values), "")


def normalize_lookup_token(value: str) -> str:
    return re.sub(r"[\s:：/\\\-_.|*（）()]+", "", str(value or "").lower())


def resolve_template_by_code(source_code: str, templates: list[str]) -> str:
    code = normalize_template_code(source_code)
    if not code:
        return ""
    matched: list[str] = []
    normalized_code = code.replace("-", "")
    for name in templates:
        normalized_name = re.sub(r"\s+", "", name or "").lower().replace("-", "").replace("_", "")
        if normalized_code in normalized_name:
            matched.append(name)
    if len(matched) == 1:
        return matched[0]
    return ""


def normalize_template_code(value: str) -> str:
    text = re.sub(r"\s+", "", value or "").lower()
    match = re.search(r"(?:r[-_ ]?)?(\d{3}[a-z])", text, flags=re.IGNORECASE)
    if not match:
        return ""
    return f"r-{match.group(1).lower()}"


def sanitize_file_name(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", value or "")
    cleaned = cleaned.strip().strip(".")
    return cleaned[:80] or "row"


def resolve_blank_template(templates: list[str]) -> str:
    exact = next((name for name in templates if name == "R-802B 空白.docx"), "")
    if exact:
        return exact
    for name in templates:
        normalized = normalize_header(name).replace("_", "-")
        if "r802b" in normalized and "空白" in name:
            return name
    for name in templates:
        if "空白" in name:
            return name
    return ""


def is_placeholder_record_row(fields: dict[str, str]) -> bool:
    device_name = normalize_header(fields.get("device_name", ""))
    device_model = normalize_header(fields.get("device_model", ""))
    device_code = normalize_header(fields.get("device_code", ""))

    name_is_placeholder = device_name in {
        "instrumentname",
        "devicename",
        "equipmentname",
        "器具名称",
        "设备名称",
        "仪器名称",
    }
    model_code_text = f"{device_model}\n{device_code}"
    model_or_code_is_placeholder = any(
        token and token in model_code_text
        for token in (
            "modelspecification",
            "model/specification",
            "规格model/specification",
            "instrumentserialnumber",
            "serialnumber",
            "型号规格",
            "型号/编号",
        )
    )
    return name_is_placeholder and model_or_code_is_placeholder
